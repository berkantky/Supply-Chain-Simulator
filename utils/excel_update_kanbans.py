"""
Ergänze Excel-Sheet mit Kanban-Anzahlen basierend auf Buffer-Zonen.
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

excel_path = Path(__file__).parent.parent / 'data' / 'raw' / 'DDSCM.xlsx'

# Lade Excel
df = pd.read_excel(excel_path, sheet_name='Pufferzonen', engine='openpyxl', header=None)

# Extrahiere Komponenten, Zonen
komponenten = df.iloc[1, 2:14].tolist()
green_zone = df.iloc[15, 2:14].astype(int).tolist()
yellow_zone = df.iloc[19, 2:14].astype(int).tolist()
red_zone = df.iloc[18, 2:14].astype(int).tolist()

# Berechne Kanban-Anzahl (Total Buffer Capacity)
kanban_count = [int(g + y + r) for g, y, r in zip(green_zone, yellow_zone, red_zone)]

# Markiere Event-Kanbans (CT7, TT8)
event_kanban_idx = []
for i, name in enumerate(komponenten):
    if str(name).strip() in ['CT7', 'TT8']:
        event_kanban_idx.append(i)

print("=" * 70)
print("KANBAN-BERECHNUNG - Komponenten mit Kanban-Anzahl:")
print("=" * 70)
for i, (name, count) in enumerate(zip(komponenten, kanban_count)):
    event_marker = " [EVENT-KANBAN]" if i in event_kanban_idx else ""
    print(f"Spalte {i+3:2d}: {str(name):25} → {count:3d} Kanbans{event_marker}")

# Schreibe mit openpyxl zurück
wb = load_workbook(excel_path)
ws = wb['Pufferzonen']

# Füge Zeile ein für Kanban-Anzahl (Row 22)
ws['A22'] = 'Number of Kanbans'
ws['B22'] = 'Anzahl Kanbans im Umlauf'

for col_idx, count in enumerate(kanban_count, start=3):  # Start bei Spalte C (3)
    cell = ws.cell(row=22, column=col_idx)
    cell.value = count
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    
    # Event-Kanbans: Orange, Normal: Blau
    if col_idx - 3 in event_kanban_idx:
        cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    else:
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    cell.alignment = Alignment(horizontal="center", vertical="center")

wb.save(excel_path)
print("\n✓ Excel aktualisiert: Row 22 mit Kanban-Anzahlen")
print(f"✓ Event-Kanbans markiert: {len(event_kanban_idx)} Produkte")
